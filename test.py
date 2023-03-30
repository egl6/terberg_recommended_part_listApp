import streamlit as st
import os
from tempfile import NamedTemporaryFile
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from io import BytesIO
import base64

st.set_page_config(page_title = 'BETA - TERBERG', layout='wide', page_icon='./terberg.png')

st.title("BETA - Recommended Part List and Filter List For Your Fleet")

tabs = st.tabs(["Welcome","Reccomended Parts","Filter List"])

tab_welcome = tabs[0]
tab_rec_parts = tabs[1]
tab_rec_filters = tabs[2]

modelsp = ['YT193-08','YT223-27','RT283-12','RT403-17']
modelsf = ['YT193-08 ','YT223-27 ','RT283-12 ','RT403-17 ']

# Create webpage
# Main page
with tab_welcome:
    st.header("Thank you for accesing this portal. Please serve yourself and download the list of parts and filters according to your fleet.")
    st.markdown("Click on tabs 'Recommended Part' and/or 'Recommended Filters', select your fleet and that's it! Download you parts list!" )

with tab_rec_parts:
    st.title("Recommended Part List For Your Fleet")
    st.header('Please Select your models:')

    if 'models_p' not in st.session_state.keys():
        models_p = modelsp
        st.session_state['models_p'] = models_p
    else:
        models_p = st.session_state['models_p']

    for i in models_p:
        st.checkbox(i, key=i)

    #st.write( [i.replace('dynamic_checkbox_','') for i in st.session_state.keys() if i.startswith('dynamic_checkbox_') and st.session_state[i]])

    files_name_list = []
    qty_per_model = {}

    # Saving Stocking coefficient
    qty_per_model['stock_coef'] = st.number_input('Stocking coeficient from 1% to 100%:', min_value=1, max_value=100, step=1, value=15)

    # If box checked qty of units will be asked
    for i in st.session_state['models_p']:
        if st.session_state[i] == True:
            var = st.number_input('Number of '+i+' units:', min_value=1, step=1)
            qty_per_model[i] = var
            excel_name = i+' Recom Part List.xlsx'
            files_name_list.append(excel_name)

    #st.write(files_name_list)

    # Funtion for merge excel excel files
    def merge_excel_files(file_list):
        # Create an empty workbook to store the merged data
        wb = Workbook()
        # Create and activate cheet
        ws1 = wb.active
        ws1.title = "Recommended Fleet part list"

        # Counter for merging
        rr = 0

        # Iterate through the list of Excel files and copy parts to new excel file
        a = 0
        for file in file_list:
            ws = wb.create_sheet(file)
            selected = load_workbook('./'+file)
            selected_sheet = selected.active
            for row in selected_sheet.rows:
                #st.write(row)
                for cell in row:
                    a+=1
                    #st.write(a)
                    ws[cell.coordinate] = cell.value

            # Type the model qty into excel
            model = ws.cell(2,2).value
            subtype = ws.cell(3,2).value
            model_subtype = model+"-"+subtype
            ws.cell(5,2).value = qty_per_model[model_subtype]

            # Type the stocking coef into excel
            ws.cell(6,2).value = qty_per_model['stock_coef']

            # Multiply Value of units by stoking coef. Recom qty per fleet
            for row in range(9,ws.max_row+1):
                ws.cell(row=row,column=4).value = int(qty_per_model[model_subtype]*(qty_per_model['stock_coef']/100)*ws.cell(row=row,column=3).value)

            # Copy parts to 1st sheet - all parts
            if rr == 0:
                tt = 8
                rr +=1
            else:
                tt = 9

            for row in ws.iter_rows(min_row = tt, values_only = True):
                cell_list = []
                cell_list.append(row[0])
                ws1.append(cell_list)
        # Consolidate
        # for row in range(2, ws1.max_row+1):
        #     ws1.merge_cells(start_row=2, start_column=1, end_row=ws1.max_row+1, end_column=4)

        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        return virtual_workbook

    # Generate part list
    generate_list = st.button("Generate recommended part list")

    if generate_list:
        final_file = merge_excel_files(files_name_list)
        #st.write(final_file)
        st.success("Parts List Generated!")

        # Download button
        st.download_button("Download Recommended Parts List",
             data=final_file,
             #mime='xlsx',
             file_name="Reco_parts.xlsx")

with tab_rec_filters:
    st.title("Filters List For Your Fleet")
    st.header('Please Select your models:')
#
#     if 'models_f' not in st.session_state.keys():
#         models_f = modelsf
#         st.session_state['models_f'] = models_f
#     else:
#         models_f = st.session_state['models_f']
#
#     for i in models_f:
#         st.checkbox(i, key=i+"1")
